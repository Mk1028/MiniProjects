import openpyxl
import click
from bs4 import BeautifulSoup

import os
import django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'IPL.settings')
django.setup()

from IPLApp.models import *

@click.group()
def main():
    print("Welcome To MySQL")

@main.command()
def importdata():
    '''Import(s) data from Matches,Deliveries sheets to Database'''
    path = 'matches.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        l = []
        matches_obj = Matches()
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            l.append(cell_obj.value)

        matches_obj.id = l[0]
        matches_obj.season = l[1]
        matches_obj.city = l[2]
        matches_obj.date = l[3]
        matches_obj.team1 = l[4]
        matches_obj.team2 = l[5]
        matches_obj.toss_winner = l[6]
        matches_obj.toss_result = l[7]
        matches_obj.result = l[8]
        matches_obj.dl_applied = l[9]
        matches_obj.winner = l[10]
        matches_obj.win_by_runs = l[11]
        matches_obj.win_by_wickets = l[12]
        matches_obj.player_of_match = l[13]
        matches_obj.venue = l[14]
        matches_obj.umpire1 = l[15]
        matches_obj.umpire2 = l[16]


        matches_obj.save()
        del matches_obj

    path = 'deliveries.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row

    for i in range(2, max_row + 1):
        l = []
        deliveries_obj = Deliveries()
        for j in range(1, max_col + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            l.append(cell_obj.value)

        deliveries_obj.dl_applied=l[1]
        deliveries_obj.batting_team = l[2]
        deliveries_obj.bowling_team = l[3]
        deliveries_obj.over = l[4]
        deliveries_obj.ball = l[5]
        deliveries_obj.batsman = l[6]
        deliveries_obj.non_striker = l[7]
        deliveries_obj.bowler = l[8]
        deliveries_obj.is_super_over = l[9]
        deliveries_obj.wide_runs = l[10]
        deliveries_obj.bye_runs = l[11]
        deliveries_obj.legbye_runs = l[12]
        deliveries_obj.noball_runs = l[13]
        deliveries_obj.penalty_runs = l[14]
        deliveries_obj.batsman_runs = l[15]
        deliveries_obj.extra_runs = l[16]
        deliveries_obj.total_runs = l[17]
        deliveries_obj.player_dismissed = l[18]
        deliveries_obj.dismissal_kind = l[19]
        deliveries_obj.fielder = l[20]

        try:
            deliveries_obj.match=Matches.objects.get(id=l[0])
            deliveries_obj.save()
        except:
            print(l[0]+" is not in College Table")
        del deliveries_obj


if __name__ == '__main__':
    main()
